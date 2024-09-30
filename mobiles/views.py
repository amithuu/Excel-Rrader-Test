import openpyxl
from rest_framework import status
from rest_framework.decorators import api_view
from rest_framework.response import Response
from .serializers import DynamicExcelDataSerializer

EXCEL_FILE_PATH = r'C:\Users\Amith\OneDrive\Pictures\APi_test\amazon\mobiles\files\testfile.xlsx'


@api_view(['GET'])
def fetch_excel_data(request):
    workbook = openpyxl.load_workbook(EXCEL_FILE_PATH)
    sheet = workbook.active
    data = []
    headers = [cell.value for cell in sheet[1]]  # Fetch header row
    
    for row in sheet.iter_rows(min_row=2, values_only=True):
        row_data = {headers[i]: row[i] for i in range(len(headers))}
        data.append(row_data)
    
    return Response(data)

@api_view(['PUT'])
def create_new_column(request):
    column_name = request.data.get('column_name')
    if not column_name:
        return Response({'error': 'Column name is required'}, status=status.HTTP_400_BAD_REQUEST)

    workbook = openpyxl.load_workbook(EXCEL_FILE_PATH)
    sheet = workbook.active
    sheet.cell(row=1, column=sheet.max_column + 1, value=column_name)
    workbook.save(EXCEL_FILE_PATH)
    return Response({'message': 'New column created'}, status=status.HTTP_201_CREATED)

@api_view(['PATCH'])
def rename_columns(request):
    rename_mapping = request.data  # Expecting a dictionary like {'old_name': 'new_name'}
    if not rename_mapping:
        return Response({'error': 'Rename mapping is required'}, status=status.HTTP_400_BAD_REQUEST)

    workbook = openpyxl.load_workbook(EXCEL_FILE_PATH)
    sheet = workbook.active
    headers = [cell.value for cell in sheet[1]]

    for col_index, header in enumerate(headers):
        if header in rename_mapping:
            sheet.cell(row=1, column=col_index + 1, value=rename_mapping[header])

    workbook.save(EXCEL_FILE_PATH)
    return Response({'message': 'Columns renamed'}, status=status.HTTP_200_OK)

@api_view(['POST'])
def update_values(request):
    serializer = DynamicExcelDataSerializer(data=request.data)
    
    if serializer.is_valid():
        row_id = serializer.validated_data.get('data').get('id')
        new_values = {k: v for k, v in serializer.validated_data.get('data').items() if k != 'id'}

        if row_id is None or not new_values:
            return Response({'error': 'ID and new values are required'}, status=status.HTTP_400_BAD_REQUEST)

        workbook = openpyxl.load_workbook(EXCEL_FILE_PATH)
        sheet = workbook.active
        headers = [cell.value for cell in sheet[1]]

        for row in range(2, sheet.max_row + 1):
            if sheet.cell(row=row, column=1).value == row_id:  # Assuming the first column is the ID
                for col_name, value in new_values.items():
                    if col_name in headers:
                        col_index = headers.index(col_name) + 1
                        sheet.cell(row=row, column=col_index, value=value)
                break

        workbook.save(EXCEL_FILE_PATH)
        return Response({'message': 'Values updated'}, status=status.HTTP_200_OK)
    
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(['DELETE'])
def empty_excel(request):
    workbook = openpyxl.load_workbook(EXCEL_FILE_PATH)
    sheet = workbook.active

    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.value = None
            
    workbook.save(EXCEL_FILE_PATH)
    return Response({'message': 'Excel file emptied'}, status=status.HTTP_204_NO_CONTENT)
